import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import re
import time
import os
import logging

class KodyParser():

    def __init__(self):
        self._url = f'https://www.kody.su/mobile/'
        self._kody_pool = self._find_all_kode()
        self._output_file_name = 'ABC_DEF_new_base.xlsx'

    def _find_all_kode(self):
        '''
        Get all mobile codes from https://www.kody.su/mobile/
        '''
        soup = BeautifulSoup(requests.get(self._url).text, 'lxml')
        kody_pool = []
        strings = soup.find_all(string=re.compile('\d\d\d'))
        for txt in strings:
            try:
                int(txt)
            except ValueError:
                pass
            else:
                if kody_pool and int(txt) < int(kody_pool[-1]):
                    return kody_pool
                kody_pool.append(txt)
        return kody_pool

    def _get_base_of_number(self, kod):
        '''
        Input data: mobile code like '977'
        Output data: [
                        [977, [1000000, 1999999], 1xxxxxx, MTC, Липецкая область], 
                        [977, [2000000, 2999999], 2xxxxxx, Мегафон, Москва], 
                        [977, [3000000, 3999999], 3xxxxxx, Билайн, Рязанская область]
                    ]
        Return [code, [interval], mask, operator, region]
        '''
        soup = BeautifulSoup(requests.get(self._url+kod).text, 'lxml')
        # Поиск всех тэгов "tr"
        parse = soup.find_all('tr')
        base = []
        for item in parse[1:]:
            # Поиск всех тэгов "td"
            information = item.find_all('td')
            # На некоторых страницах есть пустые поля
            if information[0].text[4:] == '':
                continue
            # На некоторых страницах нетипичный формат записи
            point = re.search(r'\.\.\.', information[0].text[4:])
            if point:
                buf = []
                buf.append(kod)
                buf.append(re.split(r'\.\.\.', information[0].text[4:]))
                buf.append(information[0].text[4:])
                for inf in information[1:]:
                    buf.append(inf.text)
                continue
            f = 0
            mass = []
            # Разбивка строки по 7 символов  
            for i in range(int(len(information[0].text[4:]) / 7)):
                mass.append(information[0].text[4:][f:f+7])
                f += 7
            for element in mass:
                buf=[]
                buf.append(kod)
                el = re.sub(r'x+', r'', element)
                start = el + ('0' * (7 - len(el)))
                finish = el + ('9' * (7 - len(el)))
                buf.append([start, finish])
                buf.append(information[0].text[4:])
                for inf in information[1:]:
                    buf.append(inf.text)               
                base.append(buf)
        return base

    def _to_xls(self, array, row):
        '''
        Write array to xls
        '''
        # Проверка наличия файла и создание/присоединение, в зависимости от результата
        if not os.path.exists(self._output_file_name):
            wb = Workbook()
        else:
            wb = load_workbook(self._output_file_name)
        sheet = wb.active
        if row < 2:
            sheet['A'+str(row)] = 'ABC'
            sheet['B'+str(row)] = 'from'
            sheet['C'+str(row)] = 'to'
            sheet['D'+str(row)] = 'mask'
            sheet['E'+str(row)] = 'operator'
            sheet['F'+str(row)] = 'region'

        # Заполнить данными
        for item in array:
            row += 1
            sheet['A'+str(row)] = item[0]
            sheet['B'+str(row)] = item[1][0]
            sheet['C'+str(row)] = item[1][1]
            sheet['D'+str(row)] = item[2]
            sheet['E'+str(row)] = item[3]
            sheet['F'+str(row)] = item[4]
        
        # Сохранить файл:
        wb.save(self._output_file_name)
        return row

    def main(self):
        row = 1
        for i, kod in enumerate(self._kody_pool[:3]):
            row = self._to_xls(self._get_base_of_number(kod), row)
            print(f'{i+1} step. {kod} done')
            time.sleep(2)

if __name__ == "__main__":
    KP = KodyParser()
    KP.main()